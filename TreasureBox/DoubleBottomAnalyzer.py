import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Optional
import warnings
import os
import pandas_ta as ta #勿删
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import psycopg2
from tqdm import tqdm
import time
import logging
import configparser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

warnings.filterwarnings('ignore')

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


class Config:
    def __init__(self, config_file: str = "../config.ini"):
        self.config_file = config_file
        self._validate_config_file()
        self._load_config()
        self._ensure_directories()

    def _validate_config_file(self):
        if not os.path.exists(self.config_file):
            raise FileNotFoundError(f"配置文件未找到: {os.path.abspath(self.config_file)}")

    def _load_config(self):
        config = configparser.ConfigParser()
        config.read(self.config_file, encoding='utf-8')

        db = config['DATABASE']
        self.DB_USER = db.get('user')
        self.DB_PASSWORD = db.get('password')
        self.DB_HOST = db.get('host')
        self.DB_PORT = db.get('port')
        self.DB_NAME = db.get('db_name')

        system = config['SYSTEM']
        home_dir = system.get('HOME_DIRECTORY', '~/Downloads/CoreNews_Reports')
        self.HOME_DIRECTORY = os.path.expanduser(home_dir)
        temp_dir = system.get('TEMP_DATA_DIR', 'ShareData')
        self.TEMP_DATA_DIRECTORY = os.path.join(self.HOME_DIRECTORY, temp_dir)

        self.MAX_WORKERS = system.getint('MAX_WORKERS', fallback=15)
        self.DATA_FETCH_RETRIES = system.getint('DATA_FETCH_RETRIES', fallback=3)
        self.DATA_FETCH_DELAY = system.getint('DATA_FETCH_DELAY', fallback=5)

        self.CODE_ALIASES = {'代码': '股票代码', '证券代码': '股票代码', '股票代码': '股票代码'}
        self.NAME_ALIASES = {'名称': '股票简称', '股票名称': '股票简称', '股票简称': '股票简称', '简称': '股票简称'}
        self.PRICE_ALIASES = {'最新价': '最新价', '现价': '最新价', '当前价格': '最新价', '今收盘': '最新价',
                              '收盘': '最新价', '收盘价': '最新价'}

        self.TUSHARE_TOKEN = db.get('tushare_token')  # 如果没有配置，默认为 None
        if not self.TUSHARE_TOKEN:
            raise ValueError("配置文件中缺少 'tushare_token'，请在 [DATABASE] 节点下添加。")

        log = config['LOGGING']
        self.LOG_LEVEL = log.get('LOG_LEVEL', 'INFO')
        self.LOG_DIR = os.path.join(self.HOME_DIRECTORY, log.get('LOG_DIR', 'Logs'))

        for key, val in self.__dict__.items():
            if val is None:
                raise ValueError(f"配置项 '{key}' 未设置，请在 {self.config_file} 中检查。")

    def _ensure_directories(self):
        dirs = [self.HOME_DIRECTORY, self.TEMP_DATA_DIRECTORY, self.LOG_DIR]
        for d in dirs:
            os.makedirs(d, exist_ok=True)

    def get_db_connection_string(self) -> str:
        return f"postgresql://{self.DB_USER}:{self.DB_PASSWORD}@{self.DB_HOST}:{self.DB_PORT}/{self.DB_NAME}"


def calculate_kdj_from_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    计算KDJ指标

    """
    df = df.copy()

    # 计算RSV
    low_min = df['low'].rolling(window=9).min()
    high_max = df['high'].rolling(window=9).max()

    # 避免除零错误
    denominator = high_max - low_min
    denominator = denominator.replace(0, 1)  # 将0替换为1以避免除零

    df['RSV'] = ((df['close'] - low_min) / denominator) * 100
    df['RSV'] = df['RSV'].fillna(50)  # 用50填充NaN值

    # 计算K, D, J
    df['K'] = df['RSV'].ewm(alpha=1 / 3, adjust=False).mean()
    df['D'] = df['K'].ewm(alpha=1 / 3, adjust=False).mean()
    df['J'] = 3 * df['K'] - 2 * df['D']

    # 确保K, D, J在0-100范围内
    df['K'] = df['K'].clip(0, 100)
    df['D'] = df['D'].clip(0, 100)
    df['J'] = df['J'].clip(-50, 150)  # J值可以稍微超出0-100范围

    return df


def calculate_kdj_signal_from_df(df: pd.DataFrame) -> str:
    """
    根据单只股票的DataFrame计算严格的KDJ信号。
    此函数逻辑与 SignalManager.py 中的KDJ信号逻辑完全一致。

    Args:
        df: 包含 'close', 'high', 'low' 列的股票历史数据DataFrame

    Returns:
        str: KDJ信号描述字符串，如 "极值J线反转 (K=15.2, J=-2.1)"。若无信号则返回空字符串 ""。
    """
    if len(df) < 10:
        return ""

    # 使用 pandas_ta 计算KDJ
    df = df.copy()  # 避免修改原DataFrame
    df.ta.stoch(append=True, close='close', high='high', low='low')
    kdj_cols = [col for col in df.columns if col.startswith('STOCHk_') or col.startswith('STOCHd_')]
    if len(kdj_cols) < 2:
        return ""

    k_col = kdj_cols[0]
    d_col = kdj_cols[1]
    j_col = 'KDJ_J'
    df[j_col] = 3 * df[k_col] - 2 * df[d_col]

    # 检查最新的金叉
    kdj_cross = (df[k_col] > df[d_col]) & (df[k_col].shift(1) <= df[d_col].shift(1))
    last_row = df.iloc[-1]
    prev_row = df.iloc[-2]

    # 条件1: 极值J线反转
    if prev_row[j_col] < 0 and last_row[j_col] > 5 and kdj_cross.iloc[-1]:
        return f"极值J线反转 (K={last_row[k_col]:.1f}, J={last_row[j_col]:.1f})"

    # 条件2: 底背离金叉
    window = 10
    curr_low = df['low'].iloc[-1]
    curr_k = df[k_col].iloc[-1]
    min_k_window = df[k_col].iloc[-window:-1].min()
    min_low_window = df['low'].iloc[-window:-1].min()
    is_divergence = (curr_low <= min_low_window * 1.02) & (curr_k > min_k_window * 1.1)
    if kdj_cross.iloc[-1] and is_divergence and last_row[k_col] < 30:
        return f"底背离金叉 (K={last_row[k_col]:.1f}, J={last_row[j_col]:.1f})"

    # 计算5日均线用于后续条件
    ma5 = df['close'].rolling(window=5).mean()
    above_ma5 = df['close'] > ma5

    # 检查过去5天是否有超卖
    kd_oversold = (df[k_col] < 20) & (df[d_col] < 20)
    had_oversold = kd_oversold.iloc[-5:-1].any()

    # 条件3: 趋势确认金叉
    if had_oversold and kdj_cross.iloc[-1] and above_ma5.iloc[-1]:
        return f"趋势确认金叉 (K={last_row[k_col]:.1f}, J={last_row[j_col]:.1f})"

    # 条件4: 低位超卖金叉
    if had_oversold and kdj_cross.iloc[-1]:
        return f"低位超卖金叉 (K={last_row[k_col]:.1f}, J={last_row[j_col]:.1f})"

    # 如果以上条件都不满足，则无信号
    return ""


class StockOpportunityAnalyzer:
    """
    股票操作机会时效性分析工具
    专门分析技术信号的有效期和最佳操作窗口
    """

    def __init__(self):
        self.current_date = datetime.now()

    def calculate_opportunity_window(self, signal_date: datetime, signal_strength: str, volume_score: int,
                                     breakout_amplify: bool, valley_shrink: bool, sell_signals_count: int) -> Dict:
        """
        计算操作窗口期
        """
        days_since_signal = (self.current_date - signal_date).days

        # 基于信号强度计算有效期
        base_validity = 20  # 基础有效期20天

        # 根据信号特征调整有效期
        validity_adjustment = 0
        if signal_strength == 'strong':
            validity_adjustment += 5  # 强信号增加5天
        if volume_score >= 70:
            validity_adjustment += 3  # 高量能分数增加3天
        if breakout_amplify:
            validity_adjustment += 2  # 突破放量增加2天
        if valley_shrink:
            validity_adjustment += 2  # 底部缩量增加2天
        if sell_signals_count == 0:
            validity_adjustment += 3  # 无卖出信号增加3天
        else:
            validity_adjustment -= sell_signals_count * 2  # 每个卖出信号减少2天

        adjusted_validity = base_validity + validity_adjustment
        adjusted_validity = max(5, min(60, adjusted_validity))  # 限制在5-60天

        # 计算剩余有效时间
        remaining_time = max(0, adjusted_validity - days_since_signal)

        # 计算机会衰减率
        decay_rate = min(1.0, days_since_signal / adjusted_validity)
        opportunity_score = max(0, 100 * (1 - decay_rate))

        # 操作紧迫性评估
        urgency_level = self._assess_urgency(days_since_signal, remaining_time, adjusted_validity)

        return {
            'base_validity_days': base_validity,
            'adjusted_validity_days': adjusted_validity,
            'days_since_signal': days_since_signal,
            'remaining_validity_days': remaining_time,
            'opportunity_score': opportunity_score,
            'decay_rate': decay_rate,
            'urgency_level': urgency_level,
            'valid_until': signal_date + timedelta(days=adjusted_validity),
            'signal_age_category': self._categorize_signal_age(days_since_signal)
        }

    def _assess_urgency(self, days_since_signal: int, remaining_time: int, total_validity: int) -> str:
        """
        评估操作紧迫性
        """
        if remaining_time <= 0:
            return "已过期"
        elif remaining_time <= 3:
            return "紧急"
        elif remaining_time <= 7:
            return "较高"
        elif remaining_time <= 14:
            return "中等"
        else:
            return "较低"

    def _categorize_signal_age(self, days_since_signal: int) -> str:
        """
        分类信号年龄
        """
        if days_since_signal <= 3:
            return "新鲜信号"
        elif days_since_signal <= 7:
            return "近期信号"
        elif days_since_signal <= 14:
            return "中期信号"
        elif days_since_signal <= 30:
            return "较老信号"
        else:
            return "陈旧信号"


class MACDKDJDoubleBottomAnalyzer:
    """
    MACD+KDJ双重谷形态量化分析程序 (增强版：含量能分析和时效性分析)
    仅统计最近30个交易日内进入第二个谷的形态
    并计算当天股票价格在双谷趋势中的相对状态
    """

    def __init__(self, config: Config, max_retries: int = 3):
        """
        初始化分析器
            config: 配置对象
            max_retries: 数据库连接最大重试次数
        """
        self.config = config
        self.connection_string = config.get_db_connection_string()
        self.max_retries = max_retries
        self.conn = None
        self.lock = threading.Lock()  # 用于多线程安全的数据库连接
        self.opportunity_analyzer = StockOpportunityAnalyzer()  # 新增：时效性分析器

    def connect_database(self, force_new: bool = False):
        """
        连接数据库，支持重试机制

        Args:
            force_new: 是否强制建立新连接
        """
        if self.conn and not force_new:
            try:
                # 测试连接是否仍然有效
                self.conn.cursor().execute('SELECT 1')
                return
            except:
                # 连接失效，关闭旧连接
                try:
                    self.conn.close()
                except:
                    pass
                self.conn = None

        retry_count = 0
        while retry_count < self.max_retries:
            try:
                self.conn = psycopg2.connect(self.connection_string)
                logging.info("数据库连接成功")
                return
            except Exception as e:
                retry_count += 1
                logging.warning(f"数据库连接失败 (尝试 {retry_count}/{self.max_retries}): {e}")
                if retry_count < self.max_retries:
                    time.sleep(2)  # 等待2秒后重试
                else:
                    raise ConnectionError(f"数据库连接失败，已达到最大重试次数 {self.max_retries}")

    def safe_fetch_stock_data(self, symbol: str, start_date: str = None, end_date: str = None) -> pd.DataFrame:
        """
        安全获取股票数据，支持重试机制

        Args:
            symbol: 股票代码，如 sh603233
            start_date: 开始日期，格式 YYYY-MM-DD
            end_date: 结束日期，格式 YYYY-MM-DD

        Returns:
            包含股票数据的DataFrame
        """
        retry_count = 0
        while retry_count < self.max_retries:
            try:
                with self.lock:  # 确保线程安全
                    if not self.conn:
                        self.connect_database()

                    # 构建SQL查询 (使用 amount 替代 vol)
                    query = f"""
                    SELECT trade_date, symbol, "open", "close", high, low, amount, close_normal, adj_ratio
                    FROM stock_daily_kline
                    WHERE symbol = '{symbol}'
                    """

                    if start_date:
                        query += f" AND trade_date >= '{start_date}'"
                    if end_date:
                        query += f" AND trade_date <= '{end_date}'"

                    query += " ORDER BY trade_date ASC"

                    df = pd.read_sql_query(query, self.conn)
                    df['trade_date'] = pd.to_datetime(df['trade_date'])
                    df = df.sort_values('trade_date').reset_index(drop=True)

                    # 重命名 amount 列为 vol 以保持一致性
                    df.rename(columns={'amount': 'vol'}, inplace=True)

                    # 计算KDJ指标
                    try:
                        df = calculate_kdj_from_df(df)
                    except Exception as e:
                        logging.warning(f"KDJ计算失败，跳过: {e}")
                        # 如果无法计算KDJ，手动添加空的KDJ列
                        df['K'] = np.nan
                        df['D'] = np.nan
                        df['J'] = np.nan

                    return df
            except Exception as e:
                retry_count += 1
                logging.warning(f"获取股票 {symbol} 数据失败 (尝试 {retry_count}/{self.max_retries}): {e}")

                if retry_count < self.max_retries:
                    # 关闭当前连接并重新连接
                    try:
                        if self.conn:
                            self.conn.close()
                    except:
                        pass
                    self.conn = None
                    time.sleep(1)  # 等待1秒后重试
                else:
                    logging.error(f"获取股票 {symbol} 数据最终失败: {e}")
                    return pd.DataFrame()  # 返回空DataFrame

    def get_all_symbols(self) -> List[str]:
        """
        从数据库获取所有去重的股票代码

        Returns:
            去重后的股票代码列表
        """
        retry_count = 0
        while retry_count < self.max_retries:
            try:
                with self.lock:  # 确保线程安全
                    if not self.conn:
                        self.connect_database()

                    query = """
                    SELECT DISTINCT symbol
                    FROM stock_daily_kline
                    ORDER BY symbol
                    """

                    df = pd.read_sql_query(query, self.conn)
                    symbols = df['symbol'].tolist()

                    logging.info(f"共获取到 {len(symbols)} 个不同的股票代码")
                    return symbols
            except Exception as e:
                retry_count += 1
                logging.warning(f"获取股票代码列表失败 (尝试 {retry_count}/{self.max_retries}): {e}")

                if retry_count < self.max_retries:
                    # 关闭当前连接并重新连接
                    try:
                        if self.conn:
                            self.conn.close()
                    except:
                        pass
                    self.conn = None
                    time.sleep(1)  # 等待1秒后重试
                else:
                    logging.error(f"获取股票代码列表最终失败: {e}")
                    raise

    def get_stock_name(self, symbol: str) -> str:
        """
        根据股票代码从数据库获取股票简称

        """
        try:
            with self.lock:  # 确保线程安全
                if not self.conn:
                    self.connect_database()

            # 构建SQL查询，从stock_basic_info表获取股票名称
            query = f"""
            SELECT "name"
            FROM stock_basic_info
            WHERE symbol = '{symbol[2:]}' OR ts_code = '{symbol}'
            LIMIT 1
            """

            df = pd.read_sql_query(query, self.conn)

            if not df.empty and not pd.isna(df['name'].iloc[0]):
                return df['name'].iloc[0]
            else:
                return "未知"
        except Exception as e:
            logging.warning(f"获取股票 {symbol} 名称时出错: {e}")
            return "未知"

    def is_st_stock(self, stock_name: str) -> bool:
        """
        判断是否为ST股票（含ST前缀或st）

        Args:
            stock_name: 股票简称

        Returns:
            True表示是ST股票，False表示不是
        """
        if pd.isna(stock_name) or stock_name == "未知":
            return False  # 如果无法获取股票简称，则不视为ST股票

        # 检查是否包含ST或st
        return 'ST' in stock_name.upper()

    def calculate_ma60(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        计算60日均线

        Args:
            df: 包含股票数据的DataFrame

        Returns:
            添加了MA60列的DataFrame
        """
        df['MA60'] = df['close'].rolling(window=60).mean()
        df['MA60_slope'] = df['MA60'].diff()  # 计算MA60斜率
        return df

    def calculate_macd(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        计算MACD指标

        """
        # 计算EMA(12)和EMA(26)
        df['EMA12'] = df['close'].ewm(span=12, adjust=False).mean()
        df['EMA26'] = df['close'].ewm(span=26, adjust=False).mean()

        # 计算DIF
        df['DIF'] = df['EMA12'] - df['EMA26']

        # 计算DEA
        df['DEA'] = df['DIF'].ewm(span=9, adjust=False).mean()

        # 计算MACD柱状线 (BAR )
        df['BAR'] = (df['DIF'] - df['DEA']) * 2

        return df

    def find_local_extremes(self, series: pd.Series, order: int = 5) -> Tuple[List[int], List[int]]:
        """
        寻找局部极值点

        """
        local_minima = []
        local_maxima = []

        for i in range(order, len(series) - order):
            is_min = True
            is_max = True

            for j in range(i - order, i + order + 1):
                if j != i:
                    if series.iloc[i] > series.iloc[j]:
                        is_min = False
                    if series.iloc[i] < series.iloc[j]:
                        is_max = False

            if is_min:
                local_minima.append(i)
            elif is_max:
                local_maxima.append(i)

        return local_minima, local_maxima

    def analyze_volume(self, df: pd.DataFrame, valley1_idx: int, valley2_idx: int, convergence_idx: int) -> Dict:
        """
        分析双底形态的量能情况，验证真实性

        Args:
            df: 包含股票数据的DataFrame (含vol列)
            valley1_idx: 第一个谷的索引
            valley2_idx: 第二个谷的索引
            convergence_idx: 形态收敛/信号确认的索引

        Returns:
            量能分析结果字典
        """
        volume_analysis = {
            'is_valley_shrink': False,  # 底部是否缩量
            'is_breakout_amplify': False,  # 突破是否放量
            'volume_score': 50  # 量能综合得分 (0-100)
        }

        window = 3
        v1_start = max(0, valley1_idx - window)
        v1_end = min(len(df), valley1_idx + window)
        v2_start = max(0, valley2_idx - window)
        v2_end = min(len(df), valley2_idx + window)

        avg_vol_v1 = df['vol'][v1_start:v1_end].mean()
        avg_vol_v2 = df['vol'][v2_start:v2_end].mean()

        if avg_vol_v1 > 0 and avg_vol_v2 > 0:
            # 谷2成交量小于谷1的70%，视为缩量
            if avg_vol_v2 < avg_vol_v1 * 0.7:
                volume_analysis['is_valley_shrink'] = True
                volume_analysis['volume_score'] += 20  # 缩量是好现象，加分

        # 2. 检查颈线突破放量
        if convergence_idx < len(df) and convergence_idx > 0:
            conv_vol = df['vol'].iloc[convergence_idx]
            # 计算突破前5日均量
            pre_5d_avg_vol = df['vol'].iloc[max(0, convergence_idx - 5):convergence_idx].mean()

            if pre_5d_avg_vol > 0:
                # 突破日成交量是前5日均量的1.5倍以上，视为放量
                if conv_vol > pre_5d_avg_vol * 1.5:
                    volume_analysis['is_breakout_amplify'] = True
                    volume_analysis['volume_score'] += 30  # 放量是关键，加更多分

        # 3. 综合评分，上限100
        volume_analysis['volume_score'] = min(volume_analysis['volume_score'], 100)

        return volume_analysis

    def calculate_neckline(self, df: pd.DataFrame, valley1_idx: int, valley2_idx: int, peak_between_idx: int) -> Dict:
        """
        计算双底形态的颈线位置

        Args:
            df: 包含股票数据的DataFrame
            valley1_idx: 第一个谷的索引
            valley2_idx: 第二个谷的索引
            peak_between_idx: 两谷之间峰值的索引

        Returns:
            包含颈线信息的字典
        """
        # 获取关键点的价格
        valley1_high = df.iloc[valley1_idx]['high']
        valley2_high = df.iloc[valley2_idx]['high']
        peak_between_price = df.iloc[peak_between_idx]['high']

        # 颈线是连接两个反弹高点的直线
        neckline_price = peak_between_price

        # 计算当前股价与颈线的相对差值
        current_price = df['close'].iloc[-1]
        neck_diff = current_price - neckline_price

        return {
            'neckline_price': round(neckline_price, 2),
            'current_price': round(current_price, 2),
            'neck_diff': round(neck_diff, 2),
            'breakout_strength': '强突破' if neck_diff > 0.05 * neckline_price else (
                '弱突破' if neck_diff > 0 else '未突破')
        }

    def calculate_price_relative_status(self, df: pd.DataFrame, valley1_idx: int, valley2_idx: int,
                                        current_idx: int) -> Dict:
        """
        计算当天股票价格在双谷趋势中的相对状态

        Args:
            df: 包含股票数据的DataFrame
            valley1_idx: 第一个谷的索引
            valley2_idx: 第二个谷的索引
            current_idx: 当前日期的索引

        Returns:
            包含价格相对状态信息的字典
        """
        # 获取关键点的价格
        valley1_close = df.iloc[valley1_idx]['close']
        valley2_close = df.iloc[valley2_idx]['close']
        current_close = df.iloc[current_idx]['close']

        # 计算价格相对位置
        min_valley_price = min(valley1_close, valley2_close)
        price_range = max(valley1_close, valley2_close) - min_valley_price

        # 计算当前价格与各谷的距离
        distance_to_valley1 = current_close - valley1_close
        distance_to_valley2 = current_close - valley2_close

        # 计算当前价格与双谷平均价格的偏离
        avg_valley_price = (valley1_close + valley2_close) / 2

        # 计算价格恢复程度（相对于最低谷的恢复比例）
        recovery_rate = (current_close - min_valley_price) / (
                max(valley1_close, valley2_close) - min_valley_price) * 100

        # 判断当前价格状态
        if current_close <= min_valley_price:
            status = "低于双谷底部"
        elif current_close >= max(valley1_close, valley2_close):
            status = "高于双谷顶部"
        elif current_close >= avg_valley_price:
            status = "双谷区间上半部分"
        else:
            status = "双谷区间下半部分"

        return {
            'distance_to_valley1': round(distance_to_valley1, 2),
            'distance_to_valley2': round(distance_to_valley2, 2),
            'price_recovery_rate': round(recovery_rate, 2),
            'status_description': status,
            'current_price': round(current_close, 2),
            'valley1_price': round(valley1_close, 2),
            'valley2_price': round(valley2_close, 2)
        }

    def generate_trade_recommendation(self, signal_info: Dict) -> str:
        """
        根据信号信息生成买卖建议

        Args:
            signal_info: 信号信息字典

        Returns:
            买卖建议字符串
        """
        # 获取关键指标
        signal_strength = signal_info.get('signal_strength', 'weak')
        neckline_breakout = signal_info.get('neckline_info', {}).get('breakout_strength', '未突破') != '未突破'
        sell_signals_count = signal_info.get('sell_signals_count', 0)
        price_recovery_rate = signal_info.get('price_relative_status', {}).get('price_recovery_rate', 0)
        current_ma60_slope = signal_info.get('current_ma60_slope', 0)
        # --- 新增：获取量能分数 ---
        volume_score = signal_info.get('volume_score', 50)

        # 基础判断逻辑
        base_recommendation = '观望'
        if signal_strength == 'strong' and neckline_breakout and current_ma60_slope > 0:
            if sell_signals_count == 0:
                base_recommendation = '强烈买入'
            elif sell_signals_count == 1:
                base_recommendation = '谨慎买入'
            else:
                base_recommendation = '观望'
        elif signal_strength == 'strong' and neckline_breakout:
            if sell_signals_count == 0:
                base_recommendation = '买入'
            elif sell_signals_count == 1:
                base_recommendation = '谨慎买入'
            else:
                base_recommendation = '观望'
        elif signal_strength == 'strong':
            if sell_signals_count == 0:
                base_recommendation = '潜在买入'
            else:
                base_recommendation = '观望'
        elif signal_strength == 'weak':
            if sell_signals_count == 0:
                base_recommendation = '轻仓试探'
            else:
                base_recommendation = '观望'

        # 如果基础建议是积极的，但量能不佳，则降级
        if base_recommendation in ['强烈买入', '买入', '潜在买入', '谨慎买入'] and volume_score < 50:
            if volume_score < 30:
                return '量能不足-观望'
            else:
                return '量能一般-谨慎'
        elif base_recommendation in ['强烈买入', '买入'] and volume_score < 70:
            return '量能尚可-买入'  # 介于中间的评价

        return base_recommendation

    def detect_recent_double_bottom_pattern(self, df: pd.DataFrame, recent_days: int = 30, external_kdj_signals=None) -> \
            List[Dict]:
        """
        检测最近30个交易日内进入第二个谷的MACD+KDJ双重谷形态

        Args:
            df: 包含MACD和KDJ指标的DataFrame
            recent_days: 最近交易日数量，默认30天
            external_kdj_signals: 外部传入的KDJ信号列表

        Returns:
            检测到的双重谷形态列表，每个元素包含形态详细信息
        """
        patterns = []

        # 获取BAR列
        bar_series = df['BAR']

        # 寻找局部最小值（谷底）
        minima_indices, maxima_indices = self.find_local_extremes(bar_series, order=5)

        # 过滤出负值的局部最小值（在0轴下方）
        negative_minima = [idx for idx in minima_indices if bar_series.iloc[idx] < 0]

        # 确定最近的交易日范围
        latest_date = df['trade_date'].max()
        cutoff_date = latest_date - pd.Timedelta(days=recent_days)

        # 寻找双重谷形态
        for i in range(len(negative_minima) - 1):
            valley1_idx = negative_minima[i]
            valley2_idx = negative_minima[i + 1]

            # 检查谷1和谷2之间是否有峰值
            peak_between = None
            for peak_idx in maxima_indices:
                if valley1_idx < peak_idx < valley2_idx:
                    peak_between = peak_idx
                    break

            if peak_between is None:
                continue

            # 检查谷2是否在最近30天内
            valley2_date = df.iloc[valley2_idx]['trade_date']
            if valley2_date < cutoff_date:
                continue  # 谷2不在最近30天内，跳过

            # 检查谷2是否与谷1相近（差异不超过50%）
            valley1_val = bar_series.iloc[valley1_idx]
            valley2_val = bar_series.iloc[valley2_idx]

            # 谷2不能比谷1深太多（绝对值差异相对较小）
            if abs(valley2_val) > abs(valley1_val) * 1.5:
                continue

            # 检查谷2之后是否有真正的金叉或BAR线向上穿越0轴的信号
            # 并且KDJ也显示超卖后反弹
            signal_found = False
            signal_date = None
            signal_idx = None

            # 从谷2之后开始寻找有效的买入信号
            for j in range(valley2_idx + 1, len(df) - 1):
                if j >= len(df):
                    break

                # 检查是否出现DIF上穿DEA（金叉）
                if j < len(df) - 1:  # 确保有足够的数据进行比较
                    dif_current = df['DIF'].iloc[j]
                    dif_prev = df['DIF'].iloc[j - 1] if j > 0 else df['DIF'].iloc[j]
                    dea_current = df['DEA'].iloc[j]
                    dea_prev = df['DEA'].iloc[j - 1] if j > 0 else df['DEA'].iloc[j]

                    # DIF上穿DEA形成金叉
                    if dif_prev <= dea_prev and dif_current > dea_current:
                        # 检查KDJ是否也在超卖区（<20）开始反弹
                        k_current = df['K'].iloc[j]
                        d_current = df['D'].iloc[j]
                        j_current = df['J'].iloc[j]

                        # 检查KDJ是否从超卖区（<20）开始上升
                        k_prev = df['K'].iloc[j - 1] if j > 0 else df['K'].iloc[j]
                        d_prev = df['D'].iloc[j - 1] if j > 0 else df['D'].iloc[j]

                        # KDJ超卖反弹条件
                        kdj_bullish = (k_prev <= 20 and k_current > k_prev) or \
                                      (d_prev <= 20 and d_current > d_prev) or \
                                      (k_current > d_current and k_prev <= d_prev)  # K上穿D

                        # 或者KDJ从低位开始向上
                        kdj_low_and_rising = (k_current < 30 and k_current > k_prev) and \
                                             (d_current < 30 and d_current > d_prev)

                        # 检查是否存在高位死叉信号（作为风险警示）
                        has_high_sell_signal = any(
                            sig['signal_type'] == 'sell' and sig['date'] == df.iloc[j]['trade_date'].date()
                            for sig in external_kdj_signals or self.detect_kdj_signals(df.loc[:j])
                        )

                        # 检查MA60方向是否向上
                        ma60_slope = df['MA60_slope'].iloc[j] if j < len(df) else 0

                        # 只有当MACD金叉 + KDJ共振 + MA60方向向上时才确认信号
                        # 同时记录是否存在高位死叉信号
                        if (kdj_bullish or kdj_low_and_rising) and ma60_slope > 0:
                            signal_found = True
                            signal_idx = j
                            signal_date = df.iloc[j]['trade_date'].date()
                            break

                    # 或者BAR由负转正
                    bar_current = bar_series.iloc[j]
                    bar_prev = bar_series.iloc[j - 1] if j > 0 else bar_series.iloc[j]

                    if bar_prev <= 0 and bar_current > 0:
                        # 检查KDJ是否也在超卖区反弹
                        k_current = df['K'].iloc[j]
                        d_current = df['D'].iloc[j]
                        j_current = df['J'].iloc[j]

                        # 检查KDJ是否从超卖区（<20）开始上升
                        k_prev = df['K'].iloc[j - 1] if j > 0 else df['K'].iloc[j]
                        d_prev = df['D'].iloc[j - 1] if j > 0 else df['D'].iloc[j]

                        # KDJ超卖反弹条件
                        kdj_bullish = (k_prev <= 20 and k_current > k_prev) or \
                                      (d_prev <= 20 and d_current > d_prev) or \
                                      (k_current > d_current and k_prev <= d_prev)  # K上穿D

                        # 或者KDJ从低位开始向上
                        kdj_low_and_rising = (k_current < 30 and k_current > k_prev) and \
                                             (d_current < 30 and d_current > d_prev)

                        # 检查是否存在高位死叉信号（作为风险警示）
                        has_high_sell_signal = any(
                            sig['signal_type'] == 'sell' and sig['date'] == df.iloc[j]['trade_date'].date()
                            for sig in external_kdj_signals or self.detect_kdj_signals(df.loc[:j])
                        )

                        # 检查MA60方向是否向上
                        ma60_slope = df['MA60_slope'].iloc[j] if j < len(df) else 0

                        # 只有当BAR转正 + KDJ共振 + MA60方向向上时才确认信号
                        if (kdj_bullish or kdj_low_and_rising) and ma60_slope > 0:
                            signal_found = True
                            signal_idx = j
                            signal_date = df.iloc[j]['trade_date'].date()
                            break

            if signal_found and signal_date:
                # 计算价格相对状态
                price_status = self.calculate_price_relative_status(df, valley1_idx, valley2_idx, len(df) - 1)

                # 计算颈线信息
                neckline_info = self.calculate_neckline(df, valley1_idx, valley2_idx, peak_between)

                # 检测KDJ信号
                kdj_signals = external_kdj_signals or self.detect_kdj_signals(df, lookback_days=30)

                # 统计卖出信号数量
                sell_signals_count = sum(1 for sig in kdj_signals if sig['signal_type'] == 'sell')

                volume_analysis = self.analyze_volume(df, valley1_idx, valley2_idx, signal_idx)

                signal_datetime = datetime.combine(signal_date, datetime.min.time())
                opportunity_info = self.opportunity_analyzer.calculate_opportunity_window(
                    signal_datetime,
                    'strong',
                    volume_analysis['volume_score'],
                    volume_analysis['is_breakout_amplify'],
                    volume_analysis['is_valley_shrink'],
                    sell_signals_count
                )

                pattern_info = {
                    'valley1_date': df.iloc[valley1_idx]['trade_date'].date(),
                    'valley1_value': valley1_val,
                    'valley1_index': valley1_idx,
                    'valley1_price': df.iloc[valley1_idx]['close'],
                    'valley2_date': df.iloc[valley2_idx]['trade_date'].date(),
                    'valley2_value': valley2_val,
                    'valley2_index': valley2_idx,
                    'valley2_price': df.iloc[valley2_idx]['close'],
                    'peak_between_date': df.iloc[peak_between]['trade_date'].date(),
                    'peak_between_value': bar_series.iloc[peak_between],
                    'peak_between_idx': peak_between,
                    'convergence_date': signal_date,
                    'convergence_index': signal_idx,
                    'ma60_direction': 'up',
                    'signal_strength': 'strong',
                    'price_relative_status': price_status,
                    'neckline_info': neckline_info,
                    'kdj_signals': kdj_signals,
                    'sell_signals_count': sell_signals_count,
                    'risk_level': '高' if sell_signals_count > 0 else '低',
                    'volume_analysis': volume_analysis,
                    'volume_score': volume_analysis['volume_score'],
                    'is_valley_shrink': volume_analysis['is_valley_shrink'],
                    'is_breakout_amplify': volume_analysis['is_breakout_amplify'],
                    'opportunity_info': opportunity_info
                }

                patterns.append(pattern_info)

        return patterns

    def detect_kdj_signals(self, df: pd.DataFrame, lookback_days: int = 30) -> List[Dict]:
        """
        检测KDJ信号
        """
        # 获取最近几天的数据
        recent_df = df.tail(lookback_days).copy()

        if len(recent_df) < 10:
            return []

        # 使用KDJAnalyzer计算信号
        signal_str = calculate_kdj_signal_from_df(recent_df)

        if not signal_str:
            return []

        # 解析信号字符串
        try:
            # 分离信号类型和参数
            if '(' in signal_str and ')' in signal_str:
                signal_parts = signal_str.split('(')
                signal_type = signal_parts[0].strip()

                # 解析K和J值
                params_str = signal_parts[1].replace(')', '')
                k_val = j_val = 0

                for param in params_str.split(','):
                    param = param.strip()
                    if param.startswith('K='):
                        k_val = float(param.split('=')[1])
                    elif param.startswith('J='):
                        j_val = float(param.split('=')[1])

                # 创建信号对象
                signal_obj = {
                    'type': signal_type,
                    'date': df.iloc[-1]['trade_date'].date(),
                    'description': signal_str,
                    'strength': 'strong',
                    'signal_type': 'buy' if '金叉' in signal_type or '反转' in signal_type else 'sell',
                    'k': k_val,
                    'd': k_val,  # 在金叉情况下，D值接近K值
                    'j': j_val
                }

                return [signal_obj]
        except Exception as e:
            logging.warning(f"解析KDJ信号时出错: {e}")

        return []

    def analyze_single_symbol(self, symbol: str, start_date: str = None, end_date: str = None,
                              recent_days: int = 30) -> Dict:
        """
        分析单个股票的双重谷形态

        Args:
            symbol: 股票代码
            start_date: 开始日期
            end_date: 结束日期
            recent_days: 最近交易日数量，默认30天

        Returns:
            分析结果字典
        """
        try:
            # 获取股票简称
            stock_name = self.get_stock_name(symbol)

            # 检查是否为ST股票，如果是则直接返回空结果
            if self.is_st_stock(stock_name):
                return {
                    'symbol': symbol,
                    'stock_name': stock_name,
                    'is_st_stock': True,
                    'patterns_found': 0,
                    'patterns': [],
                    'skipped': True
                }

            # 获取数据
            df = self.safe_fetch_stock_data(symbol, start_date, end_date)

            if df.empty:
                return {'symbol': symbol, 'error': f'未找到股票 {symbol} 的数据', 'patterns_found': 0}

            # 计算指标
            df = self.calculate_ma60(df)
            df = self.calculate_macd(df)

            # 获取KDJ信号
            kdj_signals = self.detect_kdj_signals(df)
            patterns = self.detect_recent_double_bottom_pattern(df, recent_days, external_kdj_signals=kdj_signals)

            result = {
                'symbol': symbol,
                'stock_name': stock_name,
                'data_length': len(df),
                'patterns_found': len(patterns),
                'patterns': patterns,
                'latest_ma60_slope': df['MA60_slope'].iloc[-1] if len(df) > 0 else 0,
                'latest_bar_value': df['BAR'].iloc[-1] if len(df) > 0 else 0,
                'latest_close_price': df['close'].iloc[-1] if len(df) > 0 else 0,
                'analysis_period': recent_days,
                'is_st_stock': False,
                'skipped': False
            }

            return result
        except Exception as e:
            logging.error(f"分析股票 {symbol} 时发生错误: {e}")
            return {
                'symbol': symbol,
                'error': f'分析股票 {symbol} 时发生错误: {str(e)}',
                'patterns_found': 0,
                'patterns': []
            }


def export_results_to_excel(results, buy_signals, config):
    """
    将分析结果导出到Excel文件
    """
    # 创建工作簿
    wb = Workbook()

    # 删除默认工作表
    ws_main = wb.active
    ws_main.title = "双谷形态分析结果"

    # 添加表头
    headers = [
        "股票简称", "买卖建议", "信号日期", "谷1日期", "谷2日期",
        "MA60方向", "信号强度", "当前MA60斜率", "当前BAR值", "当前价格",
        "KDJ信号", "KDJ信号日期", "颈线价格", "颈线突破差值", "颈线突破强度",
        "距离谷1价格", "距离谷2价格", "谷1价格", "谷2价格",
        "量能评分", "底部缩量", "突破放量",
        "信号年龄(天)", "剩余有效期(天)", "机会得分"
    ]

    ws_main.append(headers)

    # 添加数据
    for signal in buy_signals:
        neckline_info = signal['neckline_info']
        price_status = signal['price_relative_status']
        opportunity_info = signal['opportunity_info']

        row_data = [
            signal['stock_name'],
            signal['trade_recommendation'],
            signal['signal_date'],
            signal['pattern_valley1_date'],
            signal['pattern_valley2_date'],
            signal['ma60_direction'],
            signal['signal_strength'],
            signal['current_ma60_slope'],
            signal['current_bar_value'],
            signal['current_price'],
            signal['kdj_signal'],
            signal['kdj_signal_date'],
            neckline_info['neckline_price'],
            neckline_info['neck_diff'],
            neckline_info['breakout_strength'],
            price_status['distance_to_valley1'],
            price_status['distance_to_valley2'],
            price_status['valley1_price'],
            price_status['valley2_price'],
            signal['volume_score'],
            '是' if signal['is_valley_shrink'] else '否',
            '是' if signal['is_breakout_amplify'] else '否',
            opportunity_info['days_since_signal'],
            opportunity_info['remaining_validity_days'],
            opportunity_info['opportunity_score']
        ]

        ws_main.append(row_data)

    # 设置标题样式
    header_font = Font(size=12, bold=True)
    header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 格式化表头
    for col_num, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 自动调整列宽
    for column in ws_main.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws_main.column_dimensions[column_letter].width = adjusted_width

    # 格式化数据行
    for row in ws_main.iter_rows(min_row=2, max_row=len(buy_signals) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # 保存文件
    current_time = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"双谷探底机会点_{current_time}.xlsx"
    filepath = os.path.join(config.HOME_DIRECTORY, filename)
    wb.save(filepath)
    print(f"分析结果已保存到: {filepath}")


def main():
    """
    主函数 - 全市场扫描MACD+KDJ双重谷形态（多线程版本，含时效性分析）
    """

    config = Config("../config.ini")
    # 创建分析器实例
    analyzer = MACDKDJDoubleBottomAnalyzer(config, max_retries=3)

    print("开始全市场MACD+KDJ双重谷形态扫描...")
    print("扫描条件：仅统计最近30个交易日内进入第二个谷的形态")
    print("过滤条件：MA60方向为down的形态将被排除")
    print("信号确认：等待MACD金叉/KDJ共振后再确认信号")

    # 获取所有股票代码
    all_symbols = analyzer.get_all_symbols()
    print(f"准备分析 {len(all_symbols)} 个股票...")

    # 初始化结果字典
    results = {}
    completed_count = 0
    total_count = len(all_symbols)

    # 创建锁来保护共享资源
    lock = threading.Lock()

    def update_progress():
        nonlocal completed_count
        with lock:
            completed_count += 1
            return completed_count

    def analyze_wrapper(symbol):
        result = analyzer.analyze_single_symbol(symbol)
        progress = update_progress()
        return symbol, result

    # 使用ThreadPoolExecutor进行多线程处理
    with ThreadPoolExecutor(max_workers=8) as executor:
        # 提交所有任务
        future_to_symbol = {executor.submit(analyze_wrapper, symbol): symbol for symbol in all_symbols}

        # 使用tqdm创建进度条
        with tqdm(total=len(all_symbols), desc="分析进度") as pbar:
            for future in as_completed(future_to_symbol):
                symbol, result = future.result()
                results[symbol] = result
                pbar.update(1)

                # 更新进度条描述
                completed = sum(
                    1 for r in results.values() if not r.get('is_st_stock', False) and not r.get('skipped', False))
                skipped_st = sum(1 for r in results.values() if r.get('is_st_stock', False))
                pbar.set_postfix({'有效分析': completed, '跳过ST': skipped_st})

    # 统计结果
    total_stocks = len(results)
    st_stocks = sum(1 for r in results.values() if r.get('is_st_stock', False))
    valid_stocks = total_stocks - st_stocks
    stocks_with_patterns = sum(
        1 for r in results.values() if r.get('patterns_found', 0) > 0 and not r.get('is_st_stock', False))

    print(f"\n=== 分析结果汇总 ===")
    print(f"总股票数: {total_stocks}")
    print(f"有效股票数: {valid_stocks}")

    # 生成买入信号
    buy_signals = []

    for symbol, result in results.items():

        if result.get('is_st_stock', False) or result.get('skipped', False):
            continue

        if 'patterns' in result and result['patterns']:
            for pattern in result['patterns']:
                # 获取股票简称
                stock_name = result.get('stock_name', analyzer.get_stock_name(symbol))

                # 获取最新的KDJ信号
                latest_kdj_signal = None
                latest_kdj_signal_date = None
                if pattern.get('kdj_signals'):
                    # 获取最近的KDJ信号
                    latest_kdj_signal = pattern['kdj_signals'][-1]['type']
                    latest_kdj_signal_date = pattern['kdj_signals'][-1]['date']

                # 生成买卖建议
                trade_recommendation = analyzer.generate_trade_recommendation({
                    'signal_strength': pattern['signal_strength'],
                    'neckline_info': pattern['neckline_info'],
                    'sell_signals_count': pattern['sell_signals_count'],
                    'price_relative_status': pattern['price_relative_status'],
                    'current_ma60_slope': result['latest_ma60_slope'],
                    'volume_score': pattern['volume_score']
                })

                signal = {
                    'symbol': symbol,
                    'stock_name': stock_name,
                    'signal_date': pattern['convergence_date'],
                    'pattern_valley1_date': pattern['valley1_date'],
                    'pattern_valley2_date': pattern['valley2_date'],
                    'ma60_direction': pattern['ma60_direction'],
                    'signal_strength': pattern['signal_strength'],
                    'current_ma60_slope': result['latest_ma60_slope'],
                    'current_bar_value': result['latest_bar_value'],
                    'current_price': result['latest_close_price'],
                    'kdj_signal': latest_kdj_signal or '无信号',
                    'kdj_signal_date': latest_kdj_signal_date or '无信号',
                    'price_relative_status': pattern['price_relative_status'],
                    'neckline_info': pattern['neckline_info'],
                    'sell_signals_count': pattern['sell_signals_count'],
                    'risk_level': pattern['risk_level'],
                    'has_risk_warning': pattern['sell_signals_count'] > 0,
                    'trade_recommendation': trade_recommendation,
                    'volume_score': pattern['volume_score'],
                    'is_valley_shrink': pattern['is_valley_shrink'],
                    'is_breakout_amplify': pattern['is_breakout_amplify'],
                    'opportunity_info': pattern['opportunity_info']

                }

                buy_signals.append(signal)

    print(f"总买入信号数: {len(buy_signals)}")

    if buy_signals:
        print(f"\n=== 买入信号详情 ===")

        # 按买卖建议分组
        recommendation_groups = {}
        for signal in buy_signals:
            rec = signal['trade_recommendation']
            if rec not in recommendation_groups:
                recommendation_groups[rec] = []
            recommendation_groups[rec].append(signal)

        print(f"\n按买卖建议分组:")
        for rec, signals in recommendation_groups.items():
            print(f"  {rec}: {len(signals)}个")

        for i, signal in enumerate(buy_signals):
            neckline_info = signal['neckline_info']
            price_status = signal['price_relative_status']
            opportunity_info = signal['opportunity_info']
            print(f"\n信号 {i + 1}:")
            print(f"  股票简称: {signal['stock_name']}")
            print(f"  买卖建议: {signal['trade_recommendation']}")
            print(f"  信号日期: {signal['signal_date']}")
            print(f"  信号强度: {signal['signal_strength']}")
            print(f"  MA60方向: {signal['ma60_direction']}")
            print(f"  当前价格: {signal['current_price']}")
            print(f"  KDJ信号日期: {signal['kdj_signal_date']}")
            print(f"  颈线价格: {neckline_info['neckline_price']}")
            print(f"  颈线突破差值: {neckline_info['neck_diff']}")
            print(f"  价格状态: {price_status['status_description']}")
            print(f"  风险等级: {signal['risk_level']}")
            print(f"  量能评分: {signal['volume_score']}")
            print(f"  底部缩量: {'是' if signal['is_valley_shrink'] else '否'}")
            print(f"  突破放量: {'是' if signal['is_breakout_amplify'] else '否'}")

    # 导出结果到Excel
    export_results_to_excel(results, buy_signals, config)
    print("\n分析完成!")

if __name__ == "__main__":
    main()
